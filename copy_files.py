import excel
import platform
from urllib.error import HTTPError
from urllib.error import URLError
from urllib.request import urlretrieve
from socket import error as SocketError
from urllib.request import Request, urlopen
import requests
import shutil
import os
from shutil import copyfile

import openpyxl as xl

table_location = 'excel\\a.xlsx' if platform.system() == 'Windows' else 'excel/a.xlsx'

read_urls_excel_cell_letters = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
image_file_names = []

wb = xl.load_workbook(table_location)
sh = wb['Sheet1']

for row in range(2, sh.max_row + 1):
    for column in range(2, 12):
        sh.cell(row, column)
        if sh.cell(row, column).value is not None:
            filename = sh.cell(row, column).value.split('/')[-1]
            print(f'copy row {row} column {column} :: {sh.cell(row, column).value} ')
            try:
                copyfile(sh.cell(row, column).value, 'c:/img/' + filename)
                sh.cell(row, column + 11).value = filename
            except:
                print(f'copy row {row} column {column} :: NOT FOUND')
                sh.cell(row, column + 11).value = 'NOT FOUND'
        else:
            print(f'copy row {row} column {column} :: NONE')
            sh.cell(row, column + 11).value = 'n/a'

wb.save(table_location)
# copyfile(r'C:/Users/florin.radu/Downloads/mobilepro/catalog2/product/1/0/10220088_1.jpg', 'c:/img/10220088_1.jpg')

#
# def down_loop():
#     for i in range(len(read_urls_excel_cell_letters)):
#         excel.read_image_urls(read_urls_excel_cell_letters[i])
#         download_images(image_file_names, f'_{i}')
#         excel.write_file_image_to_excel(image_file_names, image_names_cell_letters[i])
#         image_file_names = []
#
#
# down_loop()
