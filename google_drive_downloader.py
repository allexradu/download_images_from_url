from urllib.request import urlopen
import time
import requests
from selenium import webdriver
from distutils.spawn import find_executable
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
import os
import openpyxl as xl
from os import walk
from pathlib import Path
import random
import string

table_location = os.path.join(os.getcwd(), 'excel', 'a.xlsx')
wb = xl.load_workbook(table_location)
sh = wb[wb.sheetnames[0]]

IMAGES_URL_COLS = [3, 4, 5, 6, 7]
IMAGES_BLANK_COLS = [8, 9, 10, 11, 12]

image_file_names = []
test_list = ['https://drive.google.com/file/d/1SbNb34NZNGm6K8yXrkO-ePNxQ6NxpwbO/view?usp=sharing',
             'https://drive.google.com/file/d/1ZQq6hwHX9TY5IxK5iOWyJM9W1BBP28rc/view?usp=sharing',
             'https://drive.google.com/file/d/1sY6nnPDnsNyQNQsREOSq-XaM_0ofmwOB/view?usp=sharing']

'https://drive.google.com/u/0/uc?id=1SbNb34NZNGm6K8yXrkO-ePNxQ6NxpwbO&export=download'

chrome_options = Options()
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-setuid-sandbox")
chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("window-size=1400,2100")
chrome_options.add_argument('--disable-dev-shm-usage')

chromium = find_executable('chromium-browser')

if chromium:
    chrome_options.binary_location = chromium

chrome_options.add_argument('--browser.download.folderList=2')
chrome_options.add_argument(
    '--browser.helperApps.neverAsk.saveToDisk=application/octet-stream')
path = os.path.join(os.getcwd(), 'excel', 'temp')
prefs = {'download.default_directory': f'{path}'}
chrome_options.add_experimental_option('prefs', prefs)

# driver = webdriver.Chrome(ChromeDriverManager().install(), options = chrome_options)
driver = webdriver.Chrome(options=chrome_options, executable_path=r'C:\python\chromedriver\chromedriver.exe')

time.sleep(1)

for col in IMAGES_URL_COLS:
    for row in range(2, sh.max_row + 1):
        item = sh.cell(row, col).value

        if item is not None:
            start = item.find('https://drive.google.com/file/d/')
            end = item.find('/view?usp=sharing')
            file_id = item[32:end]
            link = 'https://drive.google.com/u/0/uc?id=' + file_id + '&export=download'

            print(f'col = {col} /{IMAGES_URL_COLS[-1]} row = {row}/{sh.max_row} {link}')

            driver.get(link)

            time.sleep(2)

            _, _, filenames = next(walk(path))

            filename_no_ext = filenames[0].split('.')[0]
            filename_ext = filenames[0].split('.')[1]
            random_string = ''.join(random.choice(string.ascii_letters) for i in range(5))
            filename = filename_no_ext + random_string + '.' + filename_ext
            Path(path, filenames[0]).rename(os.path.join(os.getcwd(), 'excel', 'images', filename))

            time.sleep(0.5)

            blank_col = IMAGES_BLANK_COLS[IMAGES_URL_COLS.index(col)]
            sh.cell(row, blank_col).value = filename

    wb.save(table_location)
